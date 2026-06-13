from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, send_file, abort, jsonify)
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
from functools import wraps
import io, os, secrets

from config import Config
from models import (db, User, Student, Faculty, LeaveType,
                    LeaveRequest, Notification)

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

# ════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════

def current_user():
    if 'user_id' in session:
        return db.session.get(User, session['user_id'])
    return None

def login_required(f):
    @wraps(f)
    def wrap(*a, **k):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*a, **k)
    return wrap

def role_required(*roles):
    def deco(f):
        @wraps(f)
        def wrap(*a, **k):
            u = current_user()
            if not u:
                return redirect(url_for('login'))
            if u.role not in roles:
                flash('Access denied for your role.', 'error')
                return redirect(url_for('dashboard'))
            return f(*a, **k)
        return wrap
    return deco

def notify(user_id, title, message):
    """Create in-app notification + send email if configured."""
    n = Notification(user_id=user_id, title=title, message=message)
    db.session.add(n)
    db.session.commit()
    # Email (optional)
    if app.config.get('MAIL_ENABLED'):
        try:
            send_email(user_id, title, message)
        except Exception as e:
            print(f"[email error] {e}")
    else:
        u = db.session.get(User, user_id)
        print(f"[NOTIFICATION → {u.email if u else user_id}] {title}: {message}")

def send_email(user_id, subject, body):
    import smtplib
    from email.mime.text import MIMEText
    u = db.session.get(User, user_id)
    if not u:
        return
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = app.config['MAIL_USERNAME']
    msg['To'] = u.email
    with smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT']) as s:
        s.starttls()
        s.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        s.send_message(msg)

@app.context_processor
def inject_globals():
    u = current_user()
    unread = 0
    if u:
        unread = Notification.query.filter_by(user_id=u.id, is_read=False).count()
    return dict(current_user=u, unread_count=unread, today=date.today().isoformat())

# ════════════════════════════════════════════
#  AUTH
# ════════════════════════════════════════════

@app.route('/', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        pw    = request.form.get('password', '')
        u = User.query.filter_by(email=email).first()
        if u and check_password_hash(u.password, pw):
            session['user_id'] = u.id
            return redirect(url_for('dashboard'))
        flash('Invalid email or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ════════════════════════════════════════════
#  DASHBOARD
# ════════════════════════════════════════════

@app.route('/dashboard')
@login_required
def dashboard():
    u = current_user()
    if u.role == 'student':
        s = u.student
        my = LeaveRequest.query.filter_by(student_id=s.id)
        stats = {
            'total':    my.count(),
            'approved': my.filter_by(status='approved').count(),
            'rejected': my.filter_by(status='rejected').count(),
            'pending':  my.filter(LeaveRequest.status.in_(['pending_faculty','faculty_approved'])).count(),
        }
        recent = my.order_by(LeaveRequest.applied_on.desc()).limit(5).all()
        return render_template('dashboard.html', stats=stats, recent=recent)

    if u.role == 'faculty':
        f = u.faculty
        advisee_ids = [s.id for s in f.advisees]
        q = LeaveRequest.query.filter(LeaveRequest.student_id.in_(advisee_ids)) if advisee_ids else LeaveRequest.query.filter(db.false())
        stats = {
            'total':    q.count(),
            'approved': q.filter_by(status='approved').count(),
            'rejected': q.filter_by(status='rejected').count(),
            'pending':  q.filter_by(status='pending_faculty').count(),
        }
        recent = q.order_by(LeaveRequest.applied_on.desc()).limit(5).all()
        return render_template('dashboard.html', stats=stats, recent=recent)

    # HOD/Admin
    q = LeaveRequest.query
    stats = {
        'total':    q.count(),
        'approved': q.filter_by(status='approved').count(),
        'rejected': q.filter_by(status='rejected').count(),
        'pending':  q.filter(LeaveRequest.status.in_(['pending_faculty','faculty_approved'])).count(),
    }
    recent = q.order_by(LeaveRequest.applied_on.desc()).limit(6).all()
    extra = {
        'students': Student.query.count(),
        'faculty':  Faculty.query.count(),
        'awaiting_hod': q.filter_by(status='faculty_approved').count(),
    }
    return render_template('dashboard.html', stats=stats, recent=recent, extra=extra)

# ════════════════════════════════════════════
#  STUDENT — APPLY LEAVE  (PAST-DATE BUG FIXED HERE)
# ════════════════════════════════════════════

@app.route('/apply', methods=['GET', 'POST'])
@role_required('student')
def apply_leave():
    u = current_user()
    types = LeaveType.query.all()

    if request.method == 'POST':
        try:
            lt_id = int(request.form['leave_type'])
            frm   = datetime.strptime(request.form['from_date'], '%Y-%m-%d').date()
            to    = datetime.strptime(request.form['to_date'], '%Y-%m-%d').date()
        except (KeyError, ValueError):
            flash('Please fill all fields with valid dates.', 'error')
            return render_template('apply_leave.html', types=types)

        reason = request.form.get('reason', '').strip()
        today  = date.today()

        # ─── VALIDATION (the fix) ───
        if frm < today:
            flash('You cannot apply for leave on a past date. Please select today or a future date.', 'error')
            return render_template('apply_leave.html', types=types)
        if to < frm:
            flash('End date cannot be before the start date.', 'error')
            return render_template('apply_leave.html', types=types)
        if not reason:
            flash('Please provide a reason for your leave.', 'error')
            return render_template('apply_leave.html', types=types)

        # ─── OVERLAP CHECK ───
        # Block applying if these dates clash with an existing active request
        # (pending or approved). Cancelled/rejected leaves are ignored.
        active_statuses = ['pending_faculty', 'faculty_approved', 'approved']
        clash = LeaveRequest.query.filter(
            LeaveRequest.student_id == u.student.id,
            LeaveRequest.status.in_(active_statuses),
            LeaveRequest.from_date <= to,
            LeaveRequest.to_date >= frm
        ).first()
        if clash:
            flash(f'You already have a leave request for these dates '
                  f'({clash.from_date.strftime("%d %b")} to {clash.to_date.strftime("%d %b")}, '
                  f'status: {clash.status.replace("_"," ")}). '
                  f'Please cancel it first or choose different dates.', 'error')
            return render_template('apply_leave.html', types=types)

        days = (to - frm).days + 1

        leave = LeaveRequest(
            student_id=u.student.id, leave_type_id=lt_id,
            from_date=frm, to_date=to, days=days, reason=reason,
            status='pending_faculty', qr_token=secrets.token_hex(16)
        )
        db.session.add(leave)
        db.session.commit()

        # Notify class advisor
        if u.student.advisor:
            notify(u.student.advisor.user_id, 'New Leave Request',
                   f'{u.name} ({u.student.roll_no}) applied for {days} day(s) leave from {frm} to {to}.')
        flash('Leave request submitted successfully! Awaiting faculty approval.', 'success')
        return redirect(url_for('my_leaves'))

    return render_template('apply_leave.html', types=types)

@app.route('/my-leaves')
@role_required('student')
def my_leaves():
    u = current_user()
    leaves = LeaveRequest.query.filter_by(student_id=u.student.id)\
             .order_by(LeaveRequest.applied_on.desc()).all()
    return render_template('my_leaves.html', leaves=leaves)

@app.route('/cancel/<int:lid>')
@role_required('student')
def cancel_leave(lid):
    u = current_user()
    leave = db.session.get(LeaveRequest, lid)
    if not leave or leave.student_id != u.student.id:
        abort(403)
    if leave.status not in ('pending_faculty', 'faculty_approved'):
        flash('Only pending leaves can be cancelled.', 'error')
        return redirect(url_for('my_leaves'))
    leave.status = 'cancelled'
    db.session.commit()
    flash('Leave request cancelled.', 'success')
    return redirect(url_for('my_leaves'))

# ════════════════════════════════════════════
#  FACULTY — APPROVE / REJECT (stage 1)
# ════════════════════════════════════════════

@app.route('/faculty/requests')
@role_required('faculty')
def faculty_requests():
    u = current_user()
    advisee_ids = [s.id for s in u.faculty.advisees]
    pending = LeaveRequest.query.filter(
        LeaveRequest.student_id.in_(advisee_ids),
        LeaveRequest.status == 'pending_faculty'
    ).order_by(LeaveRequest.applied_on).all() if advisee_ids else []
    return render_template('faculty_requests.html', leaves=pending)

@app.route('/faculty/action/<int:lid>', methods=['POST'])
@role_required('faculty')
def faculty_action(lid):
    u = current_user()
    leave = db.session.get(LeaveRequest, lid)
    if not leave or leave.student.faculty_id != u.faculty.id:
        abort(403)
    action  = request.form.get('action')
    comment = request.form.get('comment', '').strip()
    leave.faculty_comment   = comment
    leave.faculty_action_at = datetime.utcnow()

    student_user = leave.student.user
    if action == 'approve':
        leave.status = 'faculty_approved'
        db.session.commit()
        notify(student_user.id, 'Leave Forwarded',
               f'Your leave ({leave.from_date} to {leave.to_date}) was approved by faculty and forwarded to HOD.')
        # notify HODs
        for hod in User.query.filter_by(role='hod').all():
            notify(hod.id, 'Leave Awaiting Final Approval',
                   f'{student_user.name} leave forwarded by {u.name} for final approval.')
        flash('Leave approved and forwarded to HOD.', 'success')
    else:
        leave.status = 'rejected'
        db.session.commit()
        notify(student_user.id, 'Leave Rejected',
               f'Your leave ({leave.from_date} to {leave.to_date}) was rejected by faculty. Comment: {comment or "—"}')
        flash('Leave rejected.', 'success')
    return redirect(url_for('faculty_requests'))

# ════════════════════════════════════════════
#  HOD — FINAL APPROVAL (stage 2)
# ════════════════════════════════════════════

@app.route('/hod/requests')
@role_required('hod')
def hod_requests():
    pending = LeaveRequest.query.filter_by(status='faculty_approved')\
              .order_by(LeaveRequest.faculty_action_at).all()
    return render_template('hod_requests.html', leaves=pending)

@app.route('/hod/action/<int:lid>', methods=['POST'])
@role_required('hod')
def hod_action(lid):
    leave = db.session.get(LeaveRequest, lid)
    if not leave:
        abort(404)
    action  = request.form.get('action')
    comment = request.form.get('comment', '').strip()
    leave.hod_comment   = comment
    leave.hod_action_at = datetime.utcnow()
    student_user = leave.student.user

    if action == 'approve':
        leave.status = 'approved'
        # Attendance integration: mark approved leave days
        s = leave.student
        # (kept simple — approved leave does not reduce attendance %)
        db.session.commit()
        notify(student_user.id, 'Leave Approved ✅',
               f'Your leave ({leave.from_date} to {leave.to_date}) has been approved by HOD. QR verification available.')
        flash('Leave given final approval.', 'success')
    else:
        leave.status = 'rejected'
        db.session.commit()
        notify(student_user.id, 'Leave Rejected',
               f'Your leave ({leave.from_date} to {leave.to_date}) was rejected by HOD. Comment: {comment or "—"}')
        flash('Leave rejected.', 'success')
    return redirect(url_for('hod_requests'))

# ════════════════════════════════════════════
#  HOD/ADMIN — MANAGE USERS
# ════════════════════════════════════════════

@app.route('/manage/students', methods=['GET', 'POST'])
@role_required('hod')
def manage_students():
    if request.method == 'POST':
        name = request.form['name'].strip()
        email= request.form['email'].strip().lower()
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'error')
            return redirect(url_for('manage_students'))
        user = User(name=name, email=email,
                    password=generate_password_hash(request.form['password']),
                    role='student')
        db.session.add(user); db.session.flush()
        stu = Student(user_id=user.id, roll_no=request.form['roll_no'].strip(),
                      course=request.form['course'], semester=request.form['semester'],
                      faculty_id=int(request.form['faculty_id']) if request.form.get('faculty_id') else None)
        db.session.add(stu); db.session.commit()
        flash(f'Student {name} added.', 'success')
        return redirect(url_for('manage_students'))

    students = Student.query.join(User).order_by(User.name).all()
    faculty  = Faculty.query.join(User).all()
    return render_template('manage_students.html', students=students, faculty=faculty)

@app.route('/manage/faculty', methods=['GET', 'POST'])
@role_required('hod')
def manage_faculty():
    if request.method == 'POST':
        name = request.form['name'].strip()
        email= request.form['email'].strip().lower()
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'error')
            return redirect(url_for('manage_faculty'))
        user = User(name=name, email=email,
                    password=generate_password_hash(request.form['password']),
                    role='faculty')
        db.session.add(user); db.session.flush()
        fac = Faculty(user_id=user.id, department=request.form['department'],
                      designation=request.form['designation'])
        db.session.add(fac); db.session.commit()
        flash(f'Faculty {name} added.', 'success')
        return redirect(url_for('manage_faculty'))

    faculty = Faculty.query.join(User).order_by(User.name).all()
    return render_template('manage_faculty.html', faculty=faculty)

@app.route('/delete-user/<int:uid>')
@role_required('hod')
def delete_user(uid):
    user = db.session.get(User, uid)
    if user and user.role != 'hod':
        db.session.delete(user)
        db.session.commit()
        flash('User deleted.', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# ════════════════════════════════════════════
#  REPORTS + ANALYTICS
# ════════════════════════════════════════════

@app.route('/reports')
@role_required('hod', 'faculty')
def reports():
    u = current_user()
    if u.role == 'faculty':
        advisee_ids = [s.id for s in u.faculty.advisees]
        base = LeaveRequest.query.filter(LeaveRequest.student_id.in_(advisee_ids)) if advisee_ids else LeaveRequest.query.filter(db.false())
    else:
        base = LeaveRequest.query

    # By leave type
    by_type = []
    for lt in LeaveType.query.all():
        cnt = base.filter_by(leave_type_id=lt.id, status='approved').count()
        by_type.append({'name': lt.name, 'count': cnt})

    # Monthly (current year)
    monthly = {}
    for lr in base.filter_by(status='approved').all():
        key = lr.from_date.strftime('%b %Y')
        monthly[key] = monthly.get(key, 0) + lr.days

    # Student-wise
    student_rows = []
    students = Student.query.all() if u.role == 'hod' else u.faculty.advisees
    for s in students:
        appr = LeaveRequest.query.filter_by(student_id=s.id, status='approved').all()
        student_rows.append({
            'name': s.user.name, 'roll': s.roll_no,
            'count': len(appr), 'days': sum(l.days for l in appr)
        })

    status_counts = {
        'approved': base.filter_by(status='approved').count(),
        'rejected': base.filter_by(status='rejected').count(),
        'pending':  base.filter(LeaveRequest.status.in_(['pending_faculty','faculty_approved'])).count(),
        'cancelled':base.filter_by(status='cancelled').count(),
    }

    return render_template('reports.html', by_type=by_type, monthly=monthly,
                           student_rows=student_rows, status_counts=status_counts)

@app.route('/reports/export/excel')
@role_required('hod', 'faculty')
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    u = current_user()
    if u.role == 'faculty':
        advisee_ids = [s.id for s in u.faculty.advisees]
        leaves = LeaveRequest.query.filter(LeaveRequest.student_id.in_(advisee_ids)).all() if advisee_ids else []
    else:
        leaves = LeaveRequest.query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leave Report'
    headers = ['Student', 'Roll No', 'Course', 'Leave Type', 'From', 'To', 'Days', 'Status', 'Applied On']
    ws.append(headers)
    hdr_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
    for lr in leaves:
        ws.append([lr.student.user.name, lr.student.roll_no, lr.student.course,
                   lr.leave_type.name, lr.from_date.isoformat(), lr.to_date.isoformat(),
                   lr.days, lr.status, lr.applied_on.strftime('%Y-%m-%d')])
    for col in ws.columns:
        width = max(len(str(c.value or '')) for c in col) + 3
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='leave_report.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reports/export/pdf')
@role_required('hod', 'faculty')
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    u = current_user()
    if u.role == 'faculty':
        advisee_ids = [s.id for s in u.faculty.advisees]
        leaves = LeaveRequest.query.filter(LeaveRequest.student_id.in_(advisee_ids)).all() if advisee_ids else []
    else:
        leaves = LeaveRequest.query.all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elems = [Paragraph('College Leave Management — Leave Report', styles['Title']),
             Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", styles['Normal']),
             Spacer(1, 0.5*cm)]
    data = [['Student', 'Roll', 'Type', 'From', 'To', 'Days', 'Status']]
    for lr in leaves:
        data.append([lr.student.user.name, lr.student.roll_no, lr.leave_type.name,
                     lr.from_date.strftime('%d/%m/%y'), lr.to_date.strftime('%d/%m/%y'),
                     str(lr.days), lr.status])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f3f4f6')]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='leave_report.pdf', mimetype='application/pdf')

# ── PDF Leave Application (student downloads their own) ──
@app.route('/leave/<int:lid>/pdf')
@login_required
def leave_pdf(lid):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    u = current_user()
    leave = db.session.get(LeaveRequest, lid)
    if not leave:
        abort(404)
    # access control
    if u.role == 'student' and leave.student_id != u.student.id:
        abort(403)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle('t', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1E3A5F'))
    elems = [Paragraph('LEAVE APPLICATION', title), Spacer(1, 0.3*cm),
             Paragraph('College Leave Management System', styles['Normal']), Spacer(1, 0.6*cm)]
    info = [
        ['Student Name', leave.student.user.name],
        ['Roll Number', leave.student.roll_no],
        ['Course / Semester', f'{leave.student.course} / {leave.student.semester}'],
        ['Leave Type', leave.leave_type.name],
        ['From Date', leave.from_date.strftime('%d %B %Y')],
        ['To Date', leave.to_date.strftime('%d %B %Y')],
        ['Total Days', str(leave.days)],
        ['Reason', leave.reason],
        ['Status', leave.status.replace('_', ' ').title()],
        ['Faculty Comment', leave.faculty_comment or '—'],
        ['HOD Comment', leave.hod_comment or '—'],
        ['Applied On', leave.applied_on.strftime('%d %B %Y')],
    ]
    t = Table(info, colWidths=[5*cm, 11*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0,0), (0,-1), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
    ]))
    elems += [t, Spacer(1, 1*cm)]
    if leave.status == 'approved':
        elems.append(Paragraph(f'✓ APPROVED — Verification token: {leave.qr_token}', styles['Normal']))
    elems += [Spacer(1, 1.5*cm),
              Paragraph('_______________________<br/>Authorized Signature', styles['Normal'])]
    doc.build(elems)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'leave_application_{leave.id}.pdf', mimetype='application/pdf')

# ── QR code image + verification ──
@app.route('/leave/<int:lid>/qr')
@login_required
def leave_qr(lid):
    import qrcode
    leave = db.session.get(LeaveRequest, lid)
    if not leave:
        abort(404)
    verify_url = url_for('verify_leave', token=leave.qr_token, _external=True)
    img = qrcode.make(verify_url)
    buf = io.BytesIO()
    img.save(buf); buf.seek(0)
    return send_file(buf, mimetype='image/png')

@app.route('/verify/<token>')
def verify_leave(token):
    leave = LeaveRequest.query.filter_by(qr_token=token).first()
    return render_template('verify.html', leave=leave)

# ════════════════════════════════════════════
#  NOTIFICATIONS
# ════════════════════════════════════════════

@app.route('/notifications')
@login_required
def notifications():
    u = current_user()
    notes = Notification.query.filter_by(user_id=u.id)\
            .order_by(Notification.created_at.desc()).all()
    for n in notes:
        n.is_read = True
    db.session.commit()
    return render_template('notifications.html', notes=notes)

# ════════════════════════════════════════════
if __name__ == '__main__':
    app.run(debug=True)
